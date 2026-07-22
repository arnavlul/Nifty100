import sqlite3
import pandas as pd
import numpy as np
import sys
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '../../')))
from src.screener.engine import ScreenerEngine

def export_screener_results():
    engine = ScreenerEngine('config/screener_config.yaml')
    presets = engine.config['presets']
    
    conn = sqlite3.connect('data/nifty100.db')
    ratios = pd.read_sql_query("SELECT * FROM financial_ratios", conn)
    sectors = pd.read_sql_query("SELECT * FROM sectors", conn)
    
    market_cap_df = pd.read_sql_query("SELECT * FROM market_cap", conn)
    pl_df = pd.read_sql_query("SELECT * FROM profitandloss", conn)
    
    # We only want latest year for screening
    latest_year = ratios['year'].max()
    ratios_latest = ratios[ratios['year'] == latest_year].copy()
    mc_latest = market_cap_df[market_cap_df['year'] == latest_year].copy()
    pl_latest = pl_df[pl_df['year'] == latest_year].copy()
    
    # Merge with sectors, market_cap, and profitandloss
    df = pd.merge(ratios_latest, sectors[['company_id', 'broad_sector']], on='company_id', how='left')
    df = pd.merge(df, mc_latest[['company_id', 'pe_ratio', 'pb_ratio', 'dividend_yield_pct', 'market_cap_crore']], on='company_id', how='left')
    df = pd.merge(df, pl_latest[['company_id', 'sales']], on='company_id', how='left')
    
    df.rename(columns={
        'broad_sector': 'sector',
        'return_on_equity_pct': 'roe',
        'debt_to_equity': 'de_ratio',
        'free_cash_flow_cr': 'fcf',
        'dividend_yield_pct': 'dividend_yield',
        'market_cap_crore': 'market_cap',
        'dividend_payout_ratio_pct': 'dividend_payout',
        'interest_coverage': 'icr'
    }, inplace=True)
    
    # Mock de_declining and revenue_cagr_3yr since we didn't populate it in DB
    df['de_declining'] = False
    df['revenue_cagr_3yr'] = df.get('revenue_cagr_5yr', np.nan)
    
    wb = Workbook()
    wb.remove(wb.active) # Remove default sheet
    
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    for preset_name, rules in presets.items():
        # Apply the preset using our engine
        filtered_df = engine.apply_preset(df, preset_name)
        
        # Sort by composite_quality_score if it exists
        if 'composite_quality_score' in filtered_df.columns:
            filtered_df = filtered_df.sort_values(by='composite_quality_score', ascending=False)
            
        # Select top columns
        cols_to_display = ['company_id', 'sector', 'composite_quality_score', 'roe', 'de_ratio', 'fcf', 'revenue_cagr_5yr', 'pat_cagr_5yr', 'net_profit_margin_pct', 'roce_pct', 'cfo_pat_ratio', 'interest_coverage', 'asset_turnover', 'earnings_per_share', 'book_value_per_share', 'dividend_payout_ratio_pct']
        
        # Filter only existing columns
        cols_to_display = [c for c in cols_to_display if c in filtered_df.columns]
        
        display_df = filtered_df[cols_to_display]
        
        ws = wb.create_sheet(title=preset_name[:31]) # Excel sheet names max 31 chars
        
        for r_idx, row in enumerate(dataframe_to_rows(display_df, index=False, header=True), 1):
            ws.append(row)
            
        # Apply conditional formatting (since it's already filtered, most will be green, but we can color based on rules)
        header_row = [cell.value for cell in ws[1]]
        
        for r_idx in range(2, ws.max_row + 1):
            for c_idx, col_name in enumerate(header_row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                val = cell.value
                if val is None or isinstance(val, str):
                    continue
                    
                # Evaluate against rules to color code
                passed = None
                if col_name == 'roe' and 'roe_min' in rules:
                    passed = val > rules['roe_min']
                elif col_name == 'de_ratio' and 'de_max' in rules:
                    passed = val < rules['de_max']
                elif col_name == 'fcf' and 'fcf_min' in rules:
                    passed = val > rules['fcf_min']
                elif col_name == 'revenue_cagr_5yr' and 'revenue_cagr_5yr_min' in rules:
                    passed = val > rules['revenue_cagr_5yr_min']
                elif col_name == 'pat_cagr_5yr' and 'pat_cagr_5yr_min' in rules:
                    passed = val > rules['pat_cagr_5yr_min']
                
                if passed is True:
                    cell.fill = green_fill
                elif passed is False:
                    cell.fill = red_fill
                    
    wb.save('output/screener_output.xlsx')
    conn.close()
    print("Screener export completed successfully.")

if __name__ == '__main__':
    export_screener_results()
